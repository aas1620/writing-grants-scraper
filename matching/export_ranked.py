"""Export ranked matching results to Excel."""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import math

from scrapers.base import logger


def _s(val) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return ""
    return str(val)


OUTPUT_COLS = [
    "rank", "title", "organization", "opportunity_type", "total_score",
    "match_notes", "award_amount", "entry_fee", "deadline",
    "writing_types", "geography", "url",
]

EXCLUDED_COLS = [
    "title", "organization", "opportunity_type", "exclusion_reason",
    "writing_types", "award_amount", "deadline", "url",
]


def export_ranked(scored: list[dict], excluded: list[dict], output_path: str):
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    for i, entry in enumerate(scored, 1):
        entry["rank"] = i

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if scored:
            median_score = sorted(s["total_score"] for s in scored)[len(scored) // 2]
            top = [s for s in scored if s["total_score"] >= median_score]
            _write_sheet(writer, top, OUTPUT_COLS, "Top Matches")
            _write_sheet(writer, scored, OUTPUT_COLS, "All Scored")

        if excluded:
            _write_sheet(writer, excluded, EXCLUDED_COLS, "Excluded")

        if scored:
            def _deadline_key(x):
                d = x.get("deadline_date")
                if d and isinstance(d, str):
                    return d
                return "9999"
            by_deadline = sorted(
                [s for s in scored if _s(s.get("deadline", ""))],
                key=_deadline_key,
            )
            _write_sheet(writer, by_deadline, OUTPUT_COLS, "Calendar View")

        for sheet_name in writer.sheets:
            _format_sheet(writer.sheets[sheet_name])

    logger.info(f"Ranked results: {output_path} ({len(scored)} scored, {len(excluded)} excluded)")


def _write_sheet(writer, data: list[dict], cols: list[str], sheet_name: str):
    available_cols = [c for c in cols if any(c in d for d in data)]
    rows = [{c: d.get(c, "") for c in available_cols} for d in data]
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)


def _format_sheet(ws):
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    score_col = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "total_score":
            score_col = col_idx
            break

    if score_col:
        for row in ws.iter_rows(min_row=2, min_col=score_col, max_col=score_col):
            for cell in row:
                try:
                    val = float(cell.value)
                    if val >= 80:
                        cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
                    elif val >= 60:
                        cell.fill = PatternFill(start_color="FFEB9C", fill_type="solid")
                except (ValueError, TypeError):
                    pass
