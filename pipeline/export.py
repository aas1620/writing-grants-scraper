from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from scrapers.base import logger


TYPE_COLORS = {
    "grant": "C6EFCE",
    "fellowship": "D9E2F3",
    "residency": "FCE4D6",
    "contest": "FFF2CC",
}


def export(entries: list[dict], csv_path: str, xlsx_path: str):
    if not entries:
        logger.warning("No entries to export")
        return

    df = pd.DataFrame(entries)

    col_order = [
        "title", "organization", "opportunity_type", "writing_types",
        "award_amount", "entry_fee", "deadline", "deadline_date",
        "deadline_expired", "geography", "eligibility_notes",
        "description", "url", "source", "scraped_date",
        "id", "award_amount_numeric", "entry_fee_numeric",
    ]
    existing_cols = [c for c in col_order if c in df.columns]
    df = df[existing_cols]

    Path(csv_path).parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(csv_path, index=False)
    logger.info(f"Exported CSV: {csv_path} ({len(df)} rows)")

    _export_xlsx(df, xlsx_path)


def _export_xlsx(df: pd.DataFrame, path: str):
    Path(path).parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All", index=False)

        active = df[df["deadline_expired"] == False].copy() if "deadline_expired" in df.columns else df.copy()
        if "deadline_date" in active.columns:
            active = active.sort_values("deadline_date", na_position="last")
        active.to_excel(writer, sheet_name="Active Only", index=False)

        if "opportunity_type" in df.columns:
            for otype in sorted(df["opportunity_type"].dropna().unique()):
                sheet_name = otype.title()[:31]
                subset = df[df["opportunity_type"] == otype]
                subset.to_excel(writer, sheet_name=sheet_name, index=False)

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            _format_sheet(ws)

    logger.info(f"Exported Excel: {path} ({len(df)} rows)")


def _format_sheet(ws):
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    type_col = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "opportunity_type":
            type_col = col_idx
            break

    if type_col:
        for row in ws.iter_rows(min_row=2, min_col=type_col, max_col=type_col):
            for cell in row:
                color = TYPE_COLORS.get(str(cell.value), "")
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
